#!/usr/bin/env python3
"""Parse the Engineering papers Excel file and generate JSON + SQL for Cloudflare D1."""

import json
import re
from openpyxl import load_workbook

EXCEL_FILE = "Engineering2023-2025文章列表-引用-信息与电子工程学科-0527.xlsx"
OUTPUT_JSON = "cloudflare/src/papers.json"
OUTPUT_SQL = "cloudflare/seed.sql"

def sanitize_bibtex_key(title):
    """Generate a BibTeX key from the title."""
    # Take first 4 significant words
    words = re.sub(r'[^a-zA-Z0-9\s]', '', title).strip().split()
    key = '_'.join(w.lower() for w in words[:5])
    # Remove common stop words
    stop_words = {'the', 'a', 'an', 'of', 'for', 'in', 'to', 'and', 'on', 'with', 'by', 'from', 'at', 'is', 'it'}
    key = '_'.join(w for w in key.split('_') if w not in stop_words) or key
    if len(key) > 60:
        key = key[:60]
    return key

def parse_excel():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    papers = []
    row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if not row or not row[0]:
            continue
        
        title = str(row[0]).strip() if row[0] else ""
        paper_type = str(row[1]).strip() if row[1] else ""
        topic = str(row[2]).strip() if row[2] else ""
        year = int(row[3]) if row[3] else 0
        volume = str(row[5]).strip() if row[5] else ""
        issue = str(row[6]).strip() if row[6] else ""
        total_citations = int(row[7]) if row[7] else 0
        citations_2026 = int(row[8]) if row[8] else 0
        authors = str(row[9]).strip() if row[9] else ""
        doi = str(row[10]).strip() if row[10] else ""

        if not title:
            continue

        # Normalize topic: empty -> "未分类"
        if not topic or topic.lower() == 'none' or topic == 'N/A':
            topic = "未分类"

        bibtex_key = sanitize_bibtex_key(title)
        if doi:
            doi_clean = doi.replace("https://doi.org/", "").replace("http://dx.doi.org/", "")
        else:
            doi_clean = ""

        paper = {
            "title": title,
            "type": paper_type,
            "topic": topic,
            "year": year,
            "volume": volume,
            "issue": issue,
            "total_citations": total_citations,
            "citations_2026": citations_2026,
            "authors": authors,
            "doi": doi_clean,
            "bibtex_key": bibtex_key
        }
        papers.append(paper)

    # Save JSON
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(papers, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON saved to {OUTPUT_JSON} ({len(papers)} papers)")

    # Generate SQL
    sql_lines = []
    sql_lines.append("-- Seed data for papers table")
    sql_lines.append(f"INSERT OR IGNORE INTO users (username, password_hash) VALUES ('admin', '$2b$10$3MpYq0JcY5Xn9MWHhYbYae0qGWLQvB5G5g5J5X5Y5Z5a5b5c5d5e5f');")
    sql_lines.append("-- NOTE: Replace the password_hash above with the actual bcrypt hash of 'admin123'")
    sql_lines.append("")

    for p in papers:
        # Escape single quotes in string values
        title_esc = p['title'].replace("'", "''")
        type_esc = p['type'].replace("'", "''")
        topic_esc = p['topic'].replace("'", "''")
        authors_esc = p['authors'].replace("'", "''")
        doi_esc = p['doi'].replace("'", "''")

        sql_lines.append(
            f"INSERT OR IGNORE INTO papers (title, type, topic, year, volume, issue, total_citations, citations_2026, authors, doi, bibtex_key) "
            f"VALUES ('{title_esc}', '{type_esc}', '{topic_esc}', {p['year']}, '{p['volume']}', '{p['issue']}', "
            f"{p['total_citations']}, {p['citations_2026']}, '{authors_esc}', '{doi_esc}', '{p['bibtex_key']}');"
        )

    sql_content = "\n".join(sql_lines)
    with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
        f.write(sql_content)
    print(f"✅ SQL saved to {OUTPUT_SQL} ({len(papers)} papers)")

    # Print topic stats
    topics = {}
    for p in papers:
        t = p['topic']
        if t not in topics:
            topics[t] = {"count": 0, "total_citations": 0}
        topics[t]["count"] += 1
        topics[t]["total_citations"] += p['total_citations']
    
    print("\n📊 专题统计:")
    for t, s in sorted(topics.items(), key=lambda x: -x[1]["total_citations"]):
        print(f"  {t}: {s['count']}篇, 总引用{s['total_citations']}次")

if __name__ == "__main__":
    parse_excel()